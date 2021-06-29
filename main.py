
# python imports
import xlsxwriter 
import pandas as pd
import openpyxl as op
from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver

HEADER = [
    "URL","‪Jan 1, 2004‬","‪Feb 1, 2004‬","‪Mar 1, 2004‬","‪Apr 1, 2004‬","‪May 1, 2004‬","‪Jun 1, 2004‬","‪Jul 1, 2004‬","‪Aug 1, 2004‬","‪Sep 1, 2004‬","‪Oct 1, 2004‬",
    "‪Nov 1, 2004‬","‪Dec 1, 2004‬","‪Jan 1, 2005‬","‪Feb 1, 2005‬","‪Mar 1, 2005‬","‪Apr 1, 2005‬","‪May 1, 2005‬","‪Jun 1, 2005‬","‪Jul 1, 2005‬","‪Aug 1, 2005‬","‪Sep 1, 2005‬",
    "‪Oct 1, 2005‬","‪Nov 1, 2005‬","‪Dec 1, 2005‬","‪Jan 1, 2006‬","‪Feb 1, 2006‬","‪Mar 1, 2006‬","‪Apr 1, 2006‬","‪May 1, 2006‬","‪Jun 1, 2006‬","‪Jul 1, 2006‬","‪Aug 1, 2006‬",
    "‪Sep 1, 2006‬","‪Oct 1, 2006‬","‪Nov 1, 2006‬","‪Dec 1, 2006‬","‪Jan 1, 2007‬","‪Feb 1, 2007‬","‪Mar 1, 2007‬","‪Apr 1, 2007‬","‪May 1, 2007‬","‪Jun 1, 2007‬","‪Jul 1, 2007‬",
    "‪Aug 1, 2007‬","‪Sep 1, 2007‬","‪Oct 1, 2007‬","‪Nov 1, 2007‬","‪Dec 1, 2007‬","‪Jan 1, 2008‬","‪Feb 1, 2008‬","‪Mar 1, 2008‬","‪Apr 1, 2008‬","‪May 1, 2008‬","‪Jun 1, 2008‬",
    "‪Jul 1, 2008‬","‪Aug 1, 2008‬","‪Sep 1, 2008‬","‪Oct 1, 2008‬","‪Nov 1, 2008‬","‪Dec 1, 2008‬","‪Jan 1, 2009‬","‪Feb 1, 2009‬","‪Mar 1,2009‬","‪Apr 1, 2009‬","‪May 1, 2009‬",
    "‪Jun 1, 2009‬","‪Jul 1, 2009‬","‪Aug1, 2009‬","‪Sep 1, 2009‬","‪Oct 1, 2009‬","‪Nov 1, 2009‬","‪Dec 1, 2009‬","‪Jan 1, 2010‬","‪Feb 1, 2010‬","‪Mar 1, 2010‬","‪Apr 1, 2010‬",
    "‪May 1, 2010‬","‪Jun 1, 2010‬","‪Jul 1, 2010‬","‪Aug 1, 2010‬","‪Sep 1, 2010‬","‪Oct 1, 2010‬","‪Nov 1, 2010‬","‪Dec 1, 2010‬","‪Jan 1, 2011‬","‪Feb 1, 2011‬","‪Mar 1, 2011‬",
    "‪Apr 1, 2011‬","‪May 1, 2011‬","‪Jun 1, 2011‬","‪Jul 1, 2011‬","‪Aug 1, 2011‬","‪Sep 1, 2011‬","‪Oct 1, 2011‬","‪Nov 1, 2011‬","‪Dec 1, 2011‬","‪Jan 1, 2012‬","‪Feb 1, 2012‬",
    "‪Mar 1, 2012‬","‪Apr 1, 2012‬","‪May 1, 2012‬","‪Jun 1, 2012‬","‪Jul 1, 2012‬","‪Aug 1, 2012‬","‪Sep 1, 2012‬","‪Oct 1, 2012‬","‪Nov 1, 2012‬","‪Dec 1, 2012‬","‪Jan 1, 2013‬",
    "‪Feb 1, 2013‬","‪Mar 1, 2013‬","‪Apr 1, 2013‬","‪May 1, 2013‬","‪Jun 1, 2013‬","‪Jul 1, 2013‬","‪Aug 1, 2013‬","‪Sep 1, 2013‬","‪Oct 1, 2013‬","‪Nov 1, 2013‬","‪Dec 1, 2013‬",
    "‪Jan 1, 2014‬","‪Feb 1, 2014‬","‪Mar 1, 2014‬","‪Apr 1, 2014‬","‪May 1, 2014‬","‪Jun 1, 2014‬","‪Jul 1, 2014‬","‪Aug 1, 2014‬","‪Sep 1, 2014‬","‪Oct 1, 2014‬","‪Nov 1, 2014‬",
    "‪Dec 1, 2014‬","‪Jan 1, 2015‬","‪Feb 1, 2015‬","‪Mar 1, 2015‬","‪Apr 1, 2015‬","‪May 1, 2015‬","‪Jun 1, 2015‬","‪Jul 1, 2015‬","‪Aug 1,2015‬","‪Sep 1, 2015‬","‪Oct 1, 2015‬",
    "‪Nov 1, 2015‬","‪Dec 1, 2015‬","‪Jan1, 2016‬","‪Feb 1, 2016‬","‪Mar 1, 2016‬","‪Apr 1, 2016‬","‪May 1, 2016‬","‪Jun 1, 2016‬","‪Jul 1, 2016‬","‪Aug 1, 2016‬","‪Sep 1, 2016‬",
    "‪Oct 1, 2016‬","‪Nov 1, 2016‬","‪Dec 1, 2016‬","‪Jan 1, 2017‬","‪Feb 1, 2017‬","‪Mar 1, 2017‬","‪Apr 1, 2017‬","‪May 1, 2017‬","‪Jun 1, 2017‬","‪Jul 1, 2017‬","‪Aug 1, 2017‬",
    "‪Sep 1, 2017‬","‪Oct 1, 2017‬","‪Nov 1, 2017‬","‪Dec 1, 2017‬","‪Jan 1, 2018‬","‪Feb 1, 2018‬","‪Mar 1, 2018‬","‪Apr 1, 2018‬","‪May 1, 2018‬","‪Jun 1, 2018‬","‪Jul 1, 2018‬",
    "‪Aug 1, 2018‬","‪Sep 1, 2018‬","‪Oct 1, 2018‬","‪Nov 1, 2018‬","‪Dec 1, 2018‬","‪Jan 1, 2019‬","‪Feb 1, 2019‬","‪Mar 1, 2019‬","‪Apr 1, 2019‬","‪May 1, 2019‬","‪Jun 1, 2019‬",
    "‪Jul 1, 2019‬","‪Aug 1, 2019‬","‪Sep 1, 2019‬","‪Oct 1, 2019‬","‪Nov 1, 2019‬","‪Dec 1, 2019‬","‪Jan 1, 2020‬","‪Feb 1, 2020‬","‪Mar 1, 2020‬","‪Apr 1, 2020‬","‪May 1, 2020‬",
    "‪Jun 1, 2020‬","‪Jul 1, 2020‬","‪Aug 1, 2020‬","‪Sep 1, 2020‬","‪Oct 1, 2020‬","‪Nov 1, 2020‬","‪Dec 1, 2020‬","‪Jan 1, 2021‬","‪Feb 1, 2021‬","‪Mar 1, 2021‬","‪Apr 1, 2021‬",
    "‪May 1, 2021‬","‪Jun 1, 2021‬"
]

class google_trends:
    data = []
    source = None
    browser = None
    filename = "data.xlsx"
    headers = ['Item', 'Date', 'Count', 'URL']
    
    def __init__(self):
        self.browser = webdriver.Chrome('chromedriver.exe')
        self.browser.set_window_position(450, 30)
        self.browser.set_window_size(900, 700)

    def start(self):
        url     = 'https://trends.google.com/trends/explore?q=cars&geo=US'
        sheet   = pd.read_excel('queries.xlsx', 'Main')
        items   = sheet['Keyword']
        google  = sheet['Google URL']
        youtube = sheet['YouTube URL']
        self.create_excel_file('Google')
        self.create_excel_file('YouTube')
        self.browser.get(url)
        for index in range(1):  
            print(f'\t\t{items[index].upper()}')
            self.get_data('Google', items[index], google[index])
        # for index in range(3):  
        #     print(f'\t\t{items[index].upper()}')
        #     self.get_data('YouTube', items[index], youtube[index])
        self.browser.quit()

    def get_data(self, search, item, url):
        self.browser.get(url)
        sleep(5)
        soup    = BeautifulSoup(self.browser.page_source, features='lxml')
        table   = soup.find('table')
        try:
            tbody   = table.find('tbody')
        except:
            # Add the ITEM to MISSING List
            # missing_list.append(ITEM) 
            return
        trs     = tbody.find_all('tr')
        for tr in trs:
            value = {}
            td = tr.find_all('td')
            date = td[0].get_text()
            year = date.split(',')[1].strip(' ')
            year = int(year.replace('\u202c', ''))
            if year > 2003:
                print('"' + date + '",', end='') #, ' = ', td[1].get_text())
                value['count']  = td[1].get_text()
                value['date']   = date
                value['item']   = item
                value['url']    = url
                self.data.append(value)
                # self.write_to_excel(search, value)

    def create_excel_file(self, search):
        # creating new excle file
        workbook = xlsxwriter.Workbook(self.filename)
        # if search == 'Google':  
        workbook.add_worksheet('Google')
        # if search == 'YouTube': 
        workbook.add_worksheet('YouTube')
        workbook.close()
        workbook = op.load_workbook(self.filename, False)
        # if search == 'Google':  
        worksheet = workbook['Google']
        worksheet.append(HEADER)
        # if search == 'YouTube': 
        worksheet = workbook['YouTube']
        worksheet.append(HEADER)
        workbook.save(self.filename)
        workbook.close()

    def write_to_excel(self, search, value):
        workbook = op.load_workbook(self.filename, False)
        if search == 'Google':  worksheet = workbook['Google']
        if search == 'YouTube': worksheet = workbook['YouTube']
        items = []
        items.append(value['item'])
        items.append(value['date'])
        items.append(value['count'])
        items.append(value['url'])
        worksheet.append(items)
        workbook.save(self.filename)
        workbook.close()


if __name__ == '__main__':
    gtrend = google_trends()
    gtrend.start()


